""" 
This script will attempt to convert all information from a 
TMTool template.xlsx file into a MS TMT template file (.tb7)
"""

import os
import tkinter as tk
from lxml.etree import Element, SubElement, ElementTree
from tkinter import filedialog
import uuid
import openpyxl
import functools

# find threat from id in xlsx_threat_dict and add to .tb7
def add_element(root, ele_id, elements, _type):
    element = elements.get(ele_id)
    new_ele = Element("ElementType")
    _uuid = ''
    name = SubElement(new_ele, 'Name')
    name.text = element.get('Stencil Name')
    # check id
    id_ele = SubElement(new_ele, 'Id')
    if element.get('ID') is None:
        # generate GUID if not present
        _uuid = str(uuid.uuid4())
        id_ele.text = _uuid
    else:
        # sets to ID if present
        id_ele.text = ele_id

    desc = SubElement(new_ele, 'Description')
    desc.text = element.get('Description')

    parent = SubElement(new_ele, 'ParentElement')
    parent.text = element.get('Parent')
    
    SubElement(new_ele, 'Image')  
    desc = SubElement(new_ele, 'Hidden')
    desc.text = element.get('Hidden_bool')
    SubElement(new_ele, 'Representation')
    desc.text = element.get('Representation')

    ticc = SubElement(new_ele, 'StrokeThickness')
    ticc.text = '0'
    loc = SubElement(new_ele, 'ImageLocation')
    # default values
    loc.text = 'Centered on stencil'
    SubElement(new_ele, 'Attributes')
    SubElement(new_ele, 'StencilConstraint')
    # insert new threat
    if str(element.get('Type')) == 'GenericElements':
        root[2].insert(0, new_ele)
    elif str(element.get('Type')) == 'StandardElements':
        root[3].insert(0, new_ele)
    else:
        print('bad element type. Chose GenericElements or StandardElements')
    print('added element: ' + str(name.text) + ' ' + str(id_ele.text))
    return _uuid

#TODO: refactor
def delete_cat(root, cat):
    for item in root[4].iter():
        for subelem in item.findall('Name'):
            if subelem.text == cat:
                root[4].remove(item)
                print('removed category: ' + cat)
    return

# deletes threat based on threat_id
#TODO: refactor
def delete_threat(root, threat_id):
    for item in root[5].iter():
        for subelem in item.findall('Id'):
            if subelem.text == threat_id:
                root[5].remove(item)
                print('removed threat: ' + threat_id)
    return

# deletes element/stencil based on ele_id
#TODO: refactor
def delete_element(root, ele_id):
    for item in root[3].iter():
        for subelem in item.findall('Id'):
            if subelem.text == ele_id:
                root[3].remove(item)
                print('removed element: ' + ele_id)
    for item in root[4].iter():
        for subelem in item.findall('Id'):
            if subelem.text == ele_id:
                root[4].remove(item)
                print('removed element: ' + ele_id)
    return

# returns xlsx threats as dict of threats (dict of dicts)
def find_xlsx_threats(wb):
    sheet = wb['Threats']
    threats = dict()
    threat = dict().fromkeys({'Threat Title','Category','ID','Description', 'Include Logic', 'Exclude Logic','Properties'})
    _id = ''
    for _row in range(2,(int(sheet.max_row)+1)):
        for _col in range(1,int(sheet.max_column)):
            cell = sheet.cell(row=_row, column=_col).value
            if _col == 1:
                threat['Threat Title'] = cell
            elif _col == 2:
                threat['Category'] = cell
            elif _col == 0:
                _id = cell
                threat['ID'] = _id
            elif _col == 4:
                threat['Description'] = cell
            elif _col == 5:
                threat['Include Logic'] = cell
            elif _col == 6:
                threat['Exclude Logic'] = cell
            else:
                print("error reading xlsx!")
        # add to dict with guid as key
        threats[_id] = threat
    return threats

# returns xlsx elements as dict of elements (dict of dicts)
def find_xlsx_elements(wb):
    sheet = wb['Stencils']
    _id = ''
    elements = dict()
    element = dict().fromkeys({'Stencil Name', 'Type', 'ID', 'Description','Parent', 'Hidden_bool', 'Representation', 'Attributes'})
    for _row in range(2,(int(sheet.max_row)+1)):
            for _col in range(1,int(sheet.max_column)):
                cell = sheet.cell(row=_row, column=_col).value
                if _col == 0:
                    element['Stencil Name'] = cell
                elif _col == 1:
                    element['Type'] = cell
                elif _col == 2:
                    _id = cell
                    element['ID'] = _id
                elif _col == 3:
                    element['Description'] = cell
                elif _col == 4:
                    element['Prent'] = cell
                elif _col == 5:
                    element['Hidden_bool'] = cell
                elif _col == 6:
                    element['Representation'] = cell
                elif _col == 7:
                    element['Attributes'] = cell
                else:
                    print("error reading xlsx! 2")
            elements[_id] = element
    return elements

def get_manifest(root, ws):
    _name = ''
    _id = ''
    _ver = ''
    _author = ''
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "Template name:":
                # find Tag cell, then get the cell value that's next to the current cell
                _name = ws.cell(row=cell.row,column=(cell.column + 1)).value
            elif cell.value == "Template id:":
                _id = ws.cell(row=cell.row,column=(cell.column + 1)).value
            elif cell.value == "Template version:":
                _ver = ws.cell(row=cell.row,column=(cell.column + 1)).value
            elif cell.value == "Template author:":
                _author = ws.cell(row=cell.row,column=(cell.column + 1)).value
    SubElement(root, 'Manifest', name=_name, id=_id, version=_ver, author=_author)
    return root

# gets threat categories from the metadata worksheet
# TODO: rather with static columns; a better approach would be to find which column "Threat Categories" is in
# then search in with iter_rows(min_col= found_col, max_col=found_col). Could help avoid breaking template changes
def get_threat_categories(xml, ws):
    found = False
    root_category = SubElement(xml, 'ThreatCategories')
    for row in ws.iter_rows(max_col=1):
        for cell in row:
            if found == True:
                if cell.value is not None:
                    category = SubElement(root_category, 'ThreatCategory')
                    SubElement(category, 'Name').text = cell.value
                    # get cell next to current cell
                    SubElement(category, 'Id').text = ws.cell(row=cell.row,column=(cell.column + 1)).value
                    SubElement(category, 'ShortDescription')
                    SubElement(category, 'LongDescription')
                else:
                    return xml
            elif cell.value == "Threat Categories":
                found = True
                continue
    print("Getting threat categories failed")
    return None
 
def get_Threat_Meta(xml, ws):
    found = False
    find_list = ["Is Priority Used", "Is Status Used", "Threat Properties MetaData"]
    root_category = SubElement(xml, 'ThreatMetaData')
    props = SubElement(root_category, 'PropertiesMetaData')
    for row in ws.iter_rows(max_col=1):
        for cell in row:
            if str(cell.value) in find_list:
                if cell.value == find_list[0]:
                    SubElement(root_category, 'IsPriorityUsed').text = ws.cell(row=cell.row,column=(cell.column + 1)).value
                elif cell.value == find_list[1]:
                    SubElement(root_category, 'IsStatusUsed').text = ws.cell(row=cell.row,column=(cell.column + 1)).value
                else:
                    found = True
                    continue
            elif found:
                datum = SubElement(props, 'ThreatMetaDatum')
                SubElement(datum, 'Name').text = cell.value
                SubElement(datum, 'Label').text = ws.cell(row=cell.row,column=(cell.column + 1)).value
                SubElement(datum, 'Id').text = ws.cell(row=cell.row,column=(cell.column + 2)).value
                SubElement(datum, 'Description').text = ws.cell(row=cell.row,column=(cell.column + 3)).value
                if ws.cell(row=cell.row,column=(cell.column + 4)).value is None:
                    SubElement(datum, 'HideFromUI').text = 'false'
                else:
                    SubElement(datum, 'HideFromUI').text = 'true'
                if ws.cell(row=cell.row,column=(cell.column + 5)).value is not None:
                    SubElement(datum, 'AttributeType').text = ws.cell(row=cell.row,column=(cell.column + 4)).value
                vals = SubElement(datum, 'Values')
                val = SubElement(vals, 'Value')
    return xml

"""  <ElementType>
      <Name>Generic Interaction</Name>
      <ID>dd163aaf-713b-46df-bc66-4ace6c033067</ID>
      <Description />
      <ParentElement>ROOT</ParentElement>
      <Image>
      <Hidden>false</Hidden>
      <Representation>Ellipse</Representation>
      <StrokeThickness>0</StrokeThickness>
      <ImageLocation>Centered on stencil</ImageLocation>
      <Attributes />
      <StencilConstraints>
        <StencilConstraint>
          <SelectedStencilType>Any</SelectedStencilType>
          <SelectedStencilConnection>Any</SelectedStencilConnection>
        </StencilConstraint>
      </StencilConstraints>
    </ElementType """

# more advanced way of using openpyxl than get_metadata. This function uses find and found 
# lists to dynamically finds xlsx titles in order to avoid breaking changes
# TODO: find_list is also in template2xlsx.py as title. consolidate and refactor
# TODO: include ele_props and use offset to find starting location
def get_generic(xml, ws):
    find_list = ['Name', 'Type', 'ID', 'Description', 'ParentElement', 'Hidden', 'Representation']
    found_list = []
    llen = len(find_list)
    root_category = SubElement(xml, 'GenericElements')
    for row in ws.iter_rows(max_col=llen, max_row=1):
        for cell in row:
            found_list.insert((int(cell.column)-1),cell.value)
    delete_list = [x for x in find_list + found_list if x not in found_list]
    offset = 0
    if delete_list is not None:
        find_list = [functools.reduce(lambda item,loc: item.replace(loc,''), [item]+delete_list) for item in find_list]
        offset = len(delete_list)
        print("Error: skipping element titles that could not be found:")
        print(delete_list)
    _col = llen - offset
    for row in ws.iter_rows(max_col=_col, min_row=2):
        ele = SubElement(root_category, 'ElementType')
        for cell in row:
            title = str(ws.cell(row=1,column=cell.column).value)
            if title in find_list:
                SubElement(ele,title).text = cell.value
    return xml

def main():
    root = tk.Tk()
    root.withdraw()

    xlsx_path = None
    try:
        xlsx_path = filedialog.askopenfilename(parent=root, filetypes=[("template xlsx file", "template.xlsx")])
    except FileNotFoundError:
        print('Must choose file path, quitting... ')
        quit()
    root.destroy()
    # create 2nd file till script is g2g
    if not xlsx_path:
        print('Must choose file path, quitting... ')
        quit()
    # Open Workbook
    wb = openpyxl.load_workbook(filename=xlsx_path, data_only=True)

    # Get All Sheets
    a_sheet_names = wb.sheetnames
    metadata_sheet = wb.get_sheet_by_name(name ="Metadata")
    stencils_sheet = wb.get_sheet_by_name(name ="Stencils")
    threats_sheet = wb.get_sheet_by_name(name ="Threats")
    # check for sheets
    if ('Metadata' and 'Threats' and 'Stencils') in a_sheet_names:
        print("All Sheets found!")
    else:
        print("Error! xlxs worksheets missing")
        quit()


    class XMLNamespaces:
        xsi = 'http://www.w3.org/2001/XMLSchema-instance'
        xsd = 'http://www.w3.org/2001/XMLSchema'

    root = Element(('KnowledgeBase'), nsmap={'xsi':XMLNamespaces.xsi, 'xsd':XMLNamespaces.xsd})

    # Add xml root's subelements
    # NOTE: Do not rename the worksheet default names or default title/tags for data (anything that's in bold font)
    root = get_manifest(root, metadata_sheet)
    root = get_Threat_Meta(root,metadata_sheet)
    root = get_generic(root,stencils_sheet)
    StandardElements = SubElement(root, 'StandardElements')
    root = get_threat_categories(root, metadata_sheet)
    ThreatTypes = SubElement(root, 'ThreatTypes')

    print('Finished!')
    # copy file and rename  extension
    tb7_path = os.path.splitext(xlsx_path)[0] + '.tb7'
    #TODO: remove once it's working
    tb7_path = tb7_path.replace('.tb7','2.tb7')
    outFile = open(tb7_path, 'wb')
    et = ElementTree(root)
    et.write(outFile, xml_declaration=True, encoding='utf-8', pretty_print=True) 

if __name__ == '__main__':
    main()
