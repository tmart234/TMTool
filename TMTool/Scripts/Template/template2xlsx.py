'''
this script will pull all threat categories, threats,
threat properties, generic elements, standard elements, boundaries, and metadata
from a MS TMT template files (.tb7) after, it creates a template.xlsx file
'''

import base64
from matplotlib import style
import xlsxwriter
import openpyxl

from lxml import etree
import tkinter as tk
from tkinter import filedialog

import os
import io
import re

namespaces = {
        'http://www.w3.org/2001/XMLSchema-instance': None # skip this namespace
        }
        

# blocking
def close_wb(_wb, _wb_file):
    while True:
        try:
            _wb.save(_wb_file)
        except openpyxl.exceptions as e:
            decision = input("Exception caught in workbook.close(): %s\n"
                                "Please close the file if it is open in Excel.\n"
                                "Try to write file again? [Y/n]: " % e)
            if decision != 'n':
                continue
        break

# pull all threat Categories and their GUIDs from the XML
def find_cats(root):
    cats={}
    cat_ID=''
    cat_name=''
    for cat in root.iter('ThreatCategory'):
        for subelem in cat.findall('Name'):
            cat_name = subelem.text
        for subelem in cat.findall('Id'):
            cat_ID = subelem.text
        cats[cat_name]=cat_ID
    return cats

# take a threat category GUID and look it up in the category dict
def cat2str(cat, cats):
    for key, value in cats.items():
        if cat == value:
            return key

# returns GUID in string using regex
def getStrGUID(_s):
    guid = re.findall("[0-9a-f]{8}-[0-9a-f]{4}-[1-5][0-9a-f]{3}-[89ab][0-9a-f]{3}-[0-9a-f]{12}", _s)
    if guid:
        return guid[0]
    else:
        # MS Azure template uses non-guid IDs with periods
        if ' ' not in _s and '.' in _s:
            return _s
        else:
            return None

# take in GUID and look up the element name
def getGUIDName(_root, ele_type, _guid):
    if _guid:
        for _type in _root.findall(ele_type):
            for types in _type.iter('ElementType'):
                for subelem in types.findall('ID'):
                    ele_id = subelem.text
                    if ele_id == _guid:
                        for subelem in types.findall('Name'):
                            return subelem.text
                    else:
                        continue
    return None

# take in GUID and look up the prop name in element attribs
def getGUIDProp(_root, ele_type, _guid):
    if _guid:
        for _type in _root.findall(ele_type):
            for types in _type.iter('ElementType'):
                for subelem in types.findall('Attributes'):
                    for subelem2 in subelem.findall('Attribute'):
                        for subelem3 in subelem2.findall('Name'):
                            prop_id = subelem3.text
                            if prop_id == _guid:
                                for d_name in subelem2.findall('DisplayName'):
                                    return d_name.text
                            else:
                                continue
    return None

# finds all items in single quotes, replace any GUIDs with names
def replaceSingleQuote(_root, txt):
    items = re.findall(r"'(.*?)'", txt, re.DOTALL)
    for item in items:
        guid = getStrGUID(item)
        if guid == None:
            continue
        else:
            name = getGUIDName(_root, "GenericElements", guid)
            if not name:
                name = getGUIDName(_root, "StandardElements", guid)
                if not name:
                    print("Error: element not found")
                    continue
            txt = txt.replace(item, name)
    return txt

# finds all props within elements and replaces guid with display name
def guid2name(_root, txt):
    # the remaining guids should be prop guids
    guids = re.findall("[0-9a-f]{8}-[0-9a-f]{4}-[1-5][0-9a-f]{3}-[89ab][0-9a-f]{3}-[0-9a-f]{12}", txt)
    for guid in guids:
        prop_name = getGUIDProp(_root, "GenericElements", guid)
        if not prop_name:
            prop_name = getGUIDProp(_root, "StandardElements", guid)
            if not prop_name:    
                print("Error: element not found")
                continue
        txt = txt.replace(guid, prop_name)
    return txt

# checks if label is an existing header, adds it if not
# then writes the value to that column and row
def writeProp(label, val, _ws, row, headers, fmt):
    col = None
    if label not in headers:
        _ws.write(0, len(headers), label, fmt)
        headers.append(label)
    col = headers.index(label)
    _ws.write(row, col, val)
    return

# writes elements (stencils, flows, and boundarys)
def write_row(_root, ele_type, stencil_worksheet, headers, headder_fmt):
    _row = 1
    for _type in _root.findall(ele_type):
        for types in _type.iter('ElementType'):
            for subelem in types.findall('Name'):
                    ele_name = subelem.text
            for subelem in types.findall('ID'):
                    ele_id = subelem.text
            for subelem in types.findall('Description'):
                    ele_desc = subelem.text
            for subelem in types.findall('ParentElement'):
                    ele_parent = subelem.text
            for subelem in types.findall('Hidden'):
                    hidden = subelem.text
            for subelem in types.findall('Representation'):
                    rep = subelem.text
            for subelem in types.findall('Image'):
                img_64 = subelem.text
            imgdata = base64.b64decode(img_64)
            img = io.BytesIO(imgdata.read())
            # will not get <StrokeThickness>, <ImageLocation>, or sencil constraints
                    # get all property data (all child elements)

            my_list = [ele_name,ele_type,ele_id,ele_desc,ele_parent,hidden,rep,img]
            for col_num, data in enumerate(my_list):
                if type(data.read()) == 'bytes':
                    img = openpyxl.drawing.image.Image(img)
                    stencil_worksheet.add_image(img, str(_row + openpyxl.utils.get_column_letter(col_num)))
                else:
                    stencil_worksheet.write(_row, col_num, data)

            # write all element attribs
            for attribs in types.findall('Attributes'):
                for attrib in attribs.findall('Attribute'):
                    _label = None
                    _val = None
                    _val_list = []
                    for name in attrib.findall('DisplayName'):
                        _label = name.text
                    for vals in attrib.findall('AttributeValues'):
                        for _v in vals.iter('Value'):
                            _val_list.append(_v.text)
                    # convert list to string
                    _val = ','.join(_val_list)
                    writeProp(_label, _val, stencil_worksheet, _row, headers, headder_fmt)
            _row = _row + 1
    return

def writeElementsAndThreats(xml_root, threat_worksheet, stencil_worksheet, headder_fmt):
    title = ''
    category = ''
    threat_id = ''
    desc = ''
    include = ''
    exclude = ''
    
    # write threat headers in xlsx worksheet
    threat_headers = ['ID', 'Threat Title', 'Category', 'Description', 'Include Logic', 'Exclude Logic']
    for col_num, data in enumerate(threat_headers):
        threat_worksheet.write(0, col_num, data, headder_fmt)

    # start at row 1
    _row = 1
    for types in xml_root.iter('ThreatType'):
        # get threat's threat logic
        for gen_filters in types.findall('GenerationFilters'):
            for include_logic in gen_filters.findall('Include'):
                include = include_logic.text
                if include:
                    # replace threat logic GUIDs with names
                    # first replace GUIDs in single quotes
                    include = replaceSingleQuote(xml_root, include)
                    # then replace GUIDs that are prop names ex: "flow.<guid>"
                    include = guid2name(xml_root, include)
            for exclude_logic in gen_filters.findall('Exclude'):
                exclude = exclude_logic.text
                if exclude:
                    exclude = replaceSingleQuote(xml_root, exclude_logic.text)
                    exclude = guid2name(xml_root, exclude)
        # get elements
        for subelem in types.findall('Category'):
            category = str(cat2str(subelem.text, find_cats(xml_root)))
        for subelem in types.findall('Id'):
            threat_id = subelem.text
        for subelem in types.findall('ShortTitle'):
            # remove curley braces for xlsx output
            title = subelem.text
        for subelem in types.findall('Description'):
            desc = subelem.text

        # WRITE EACH ROW ITERATIVELY 
        my_list = [threat_id, title, category, desc, include, exclude]

        for col_num, data in enumerate(my_list):
            threat_worksheet.write(_row, col_num, data)

        # get all property data (all child elements)
        for props in types.findall('PropertiesMetaData'):
            for prop in props.findall('ThreatMetaDatum'):
                _label = None
                _val = None
                for label in prop.findall('Label'):
                    _label = label.text
                for values in prop.findall('Values'):
                    for val in values.findall('Value'):
                        _val = val.text
                if _label:
                    writeProp(_label, _val, threat_worksheet, _row, threat_headers, headder_fmt)
        # increase row
        _row = _row + 1

    # Elements headders
    stencil_headers = ['Name', 'Type', 'ID', 'Description', 'ParentElement', 'Hidden', 'Representation']
    for col_num, data in enumerate(stencil_headers):
        stencil_worksheet.write(0, col_num, data, headder_fmt)

    # write generic elements
    write_row(xml_root, "GenericElements", stencil_worksheet, stencil_headers, headder_fmt)
    write_row(xml_root, "StandardElements", stencil_worksheet, stencil_headers, headder_fmt)
    return

def writeMetadata(xml_root, meta_worksheet, cell_format):
    # get Manifest attributes
    atrib_list = ['name','id','version','author']
    for atrib in xml_root.findall('Manifest'):
        for index in range(len(atrib_list)):
            meta_worksheet.cell(row=index+1, column=1).font = cell_format
            meta_worksheet.cell(row=index+1, column=1).value = str("Template " + atrib_list[index] + ":" )
            meta_worksheet.cell(row=index+1, column=2).value = atrib.get(atrib_list[index])
    # get threat categories as a key value pair dictionary
    # TODO: get descriptions too
    new_row = len(atrib_list)+2
    template_cats = find_cats(xml_root)
    meta_worksheet.cell(row=new_row, column=1).font = cell_format
    meta_worksheet.cell(row=new_row, column=1).value = "Threat Categories"
    #meta_worksheet.cell(new_row, 0, "Threat Categories", cell_format)
    for key,val in template_cats.items():
        new_row = 1 + new_row
        meta_worksheet.cell(row=new_row, column=1).value = key
        meta_worksheet.cell(row=new_row, column=2).value = val

    # Get all threat properties available
    new_row = new_row + 1
    for props in xml_root.findall('ThreatMetaData'):
        for priority in props.findall('IsPriorityUsed'):
            new_row = new_row + 1
            meta_worksheet.cell(row=new_row, column=1).font = cell_format
            meta_worksheet.cell(row=new_row, column=1).value = "Is Priority Used"
            meta_worksheet.cell(row=new_row, column=2).value = priority.text
        for status in props.findall('IsStatusUsed'):
            new_row = new_row + 1
            meta_worksheet.cell(row=new_row, column=1).font = cell_format
            meta_worksheet.cell(row=new_row, column=1).value = "Is Status Used"
            meta_worksheet.cell(row=new_row, column=2).value = status.text
        new_row = new_row + 1
        meta_worksheet.cell(row=new_row, column=1).font = cell_format
        meta_worksheet.cell(row=new_row, column=1).value = "Threat Properties MetaData"
        for metaprops in props.findall('PropertiesMetaData'):
            for threatmeta in metaprops.findall('ThreatMetaDatum'):
                new_row = new_row + 1
                for propname in threatmeta.findall('Name'):
                    meta_worksheet.cell(row=new_row, column=1).value = propname.text
                for proplabel in threatmeta.findall('Label'):
                    meta_worksheet.cell(row=new_row, column=2).value = proplabel.text
                for id in threatmeta.findall('Id'):
                    meta_worksheet.cell(row=new_row, column=3).value = id.text
                for des in threatmeta.findall('Description'):
                    meta_worksheet.cell(row=new_row, column=4).value = des.text
                for hide in threatmeta.findall('HideFromUI'):
                    if str(hide.text) == 'true':
                        meta_worksheet.cell(row=new_row, column=5).value = 'HIDDEN'
                for _type in threatmeta.findall('AttributeType'):
                    meta_worksheet.cell(row=new_row, column=6).value = _type.text
                
                # TODO: get list of "Values"
    return

def main():
    
    root = tk.Tk()
    root.withdraw()

    file_path = None
    try:
        file_path = filedialog.askopenfilename(parent=root, filetypes=[("MS threat template files", "*.tb7")])
    except FileNotFoundError:
        print('Must choose file path, quitting... ')
        quit()
    if not file_path:
        print('Must choose file path, quitting... ')
        quit()
    
    with io.open(file_path, 'r', encoding='utf-8') as f:
        tree = etree.parse(f)
    xml_root = tree.getroot()

    # copy file and rename  extension
    wb_file = os.path.splitext(file_path)[0]
    wb_file = wb_file + '.xlsx'
    
    # Create a workbook and add worksheets
    workbook = openpyxl.Workbook()
    cell_format = openpyxl.styles.Font(bold=True)
    meta_worksheet = workbook.create_sheet('Metadata', 0)
    threat_worksheet = workbook.create_sheet('Threats', 1)
    stencil_worksheet = workbook.create_sheet('Stencils', 2)

    writeMetadata(xml_root, meta_worksheet, cell_format)
    #writeElementsAndThreats(xml_root, threat_worksheet, stencil_worksheet, cell_format)
    close_wb(workbook, wb_file)
    return

if __name__ == '__main__':
    main()